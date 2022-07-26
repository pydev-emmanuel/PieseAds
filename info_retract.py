import os
import xlsxwriter
from openpyxl import load_workbook
from bs4 import BeautifulSoup as soup


def product_data(html_file, code):
    product_details = {}
    price_bardi = None
    price_intercar = None
    price_workbook = load_workbook("C:\\Users\\HP\\Desktop\\WORK\\RULMENTI\\rulmenti_FAG\\rulmenti_FAG.xlsx")
    price_worksheet = price_workbook["Sheet1"]
    code_column = price_worksheet["A"]
    code_column_list = [code_column[x].value for x in range(len(code_column))]
    price_column = price_worksheet["B"]
    price_column_list = [price_column[x].value for x in range(len(price_column))]
    for x in code_column_list:
        if int(code) == int(x):
            price_bardi = price_column_list[code_column_list.index(x)]
    product_details["price"] = price_bardi
    oem_equivalent = []
    html = open(html_file, "r")
    contents = html.read()
    bs_content = soup(contents, "lxml")
    # price = bs_content.find(class_="quantity quantity--pricesmall productpricetoggle__gross productpricetoggle__wholesale js-product-wholesale-toggle").findNext("div").text
    # price = int(float(price))
    # price_intercar = price - (price/100)*10
    # price_intercar = int(price_intercar)
    # product_details["price"] = price_intercar
    try:
        for tag in bs_content.find_all("li", class_="refnumbers__item"):
            if tag.find('span', class_='refnumbers__manufacturer').text == "Inter Cars cross reference":
                pass
            else:
                manufacturer = tag.find('span', class_='refnumbers__manufacturer').text
                refnumber = tag.find('span', class_='refnumbers__refnumber').text
                refnumber_mod = refnumber.replace(" ", "")
                oem_equivalent.append(f"{manufacturer} - {refnumber_mod}")

    #     if "OEM part number equivalent" in tag.text:
    #         check = tag.findNext("div").ul
    #
    #     for oem_tag in check.find_all("li", class_="refnumbers__item"):
    #         print(oem_tag)

    except UnboundLocalError:
        pass
    except AttributeError:
        pass
    product_details["oem_equivalent"] = oem_equivalent
    img_column = price_worksheet["C"]
    img_column_list = [img_column[x].value for x in range(len(img_column))]
    for x in code_column_list:
        if int(code) == int(x):
            img_src = img_column_list[code_column_list.index(x)]
    product_details["img_src"] = img_src
    # try:
    #     img_html = open(f"C:\\Users\\Gh0sT\\Desktop\\WORK\\arcuri_KYB\\photo\\{code}.html", "r")
    #     img_contents = img_html.read()
    #     img_bs_content = soup(img_contents, "lxml")
    #     img_src = img_bs_content.find("img", class_="ng-star-inserted")["src"]
    #     product_details["img_src"] = img_src
    # except:
    #     product_details["img_src"] = None
    return product_details


def product_aplicatii(html_file):
    html = open(html_file, "r")
    contents = html.read()
    bs_content = soup(contents, "lxml")
    aplicatii = {
    }
    aplicatii_list = []
    motor = None
    cod_motor = None
    ani_productie = None
    kW = None
    hp = None
    ccm = None
    for car_brand_tag in bs_content.find_all("div", class_="tree__branch js-tree-trigger is-open"):
        car_brand = car_brand_tag.text
        next_tag = car_brand_tag.findNext("div")
        for leaf in next_tag.find_all("div", class_="tree__leaf js-tree-trigger is-open"):
            try:
                if leaf.findNext("div").ul["class"] == ['tree__list']:
                    pass
            except:
                name = leaf.text
                tabel = leaf.findNext("div")
                for masina in tabel.find_all("tr", class_="datatable__rowtd datatable__clickable js-clickable-row"):
                    for info in masina.find_all(class_="datatable__item"):
                        try:
                            if info.span.text == "Engine":
                                list_engine = list(info.text)
                                del list_engine[0:30]
                                engine = "".join(list_engine)
                                motor = engine.strip()
                            elif info.span.text == "Engine codes":
                                list_engine_code = list(info.text)
                                del list_engine_code[0:30]
                                engine_code = "".join(list_engine_code)
                                cod_motor = engine_code.strip()
                            elif info.span.text == "Production years":
                                production_years_list = list(info.text)
                                del production_years_list[0:30]
                                production_years = "".join(production_years_list)
                                production_years = production_years.replace("\n", "")
                                production_years = production_years.replace(" ", "")
                                ani_productie = production_years.strip()
                            elif info.span.text == "kW":
                                kw_list = list(info.text)
                                del kw_list[0:30]
                                kw = "".join(kw_list)
                                kW = f"{kw.strip()}kW"
                            elif info.span.text == "hp":
                                hp_list = list(info.text)
                                del hp_list[0:30]
                                horse = "".join(hp_list)
                                hp = f"{horse.strip()}cp"
                            elif info.span.text == "ccm":
                                ccm_list = list(info.text)
                                del ccm_list[0:30]
                                centimetricub = "".join(ccm_list)
                                ccm = f"{centimetricub.strip()}"
                        except:
                            pass
                    aplicatii_list.append([motor, cod_motor, hp, kW, ani_productie])
                aplicatii[f"{car_brand} {name}"] = aplicatii_list
                aplicatii_list = []
    return aplicatii


def descriere(aplicatii, product_details, cod_produs):
    oem_equivalent = product_details["oem_equivalent"]
    tabel_compatibilitate = []
    tabel_echivalente = []
    for key, value in aplicatii.items():
        tabel_compatibilitate.append(f"<div><b>{key}:</b></div>")
        for val in value:
            val_string = " ".join(val)
            tabel_compatibilitate.append(f"<div>{val_string}</div>")
        tabel_compatibilitate.append("<div><br></div>")
    tabel_compatibilitate = " ".join(tabel_compatibilitate)
    for ech in oem_equivalent:
        ech_string = "".join(ech)
        tabel_echivalente.append(f"<div><b>{ech_string}</b></div>")
    tabel_echivalente = " ".join(tabel_echivalente)
    descriere = f"""<h2>Set rulment roata FAG - {cod_produs}</h2><br>
        <div><br></div>
        <h3><u>Masini compatibile:</u></h3>
        {tabel_compatibilitate}
        <div><br></div>
        <div><br></div>
        <h3>Echivalente coduri original:</h3>
        {tabel_echivalente}
        """
    return descriere

adauga_excel = []
directory = "C:\\Users\\HP\\Desktop\\WORK\\RULMENTI\\rulmenti_FAG\\intercar"
for html_file in os.listdir(directory):
    cod_produs = html_file.replace(".html", "")
    print(cod_produs)
    html = f"{directory}\\{html_file}"
    product_details = product_data(html, cod_produs)
    aplicatii_produs = product_aplicatii(html)
    descriere_anunt = descriere(aplicatii_produs, product_details, cod_produs)
    oem_equivalent = product_details["oem_equivalent"]
    img_src = product_details["img_src"]
    price = product_details["price"]
    for key in aplicatii_produs.keys():
        titlu = f"Rulment roata {key}  FAG {cod_produs}"
        adauga_excel.append([titlu, "Rulmenti", descriere_anunt, "RON", price, "1", img_src])

workbook = xlsxwriter.Workbook("C:\\Users\\HP\\Desktop\\WORKBOOK\\rulmenti_FAG.xlsx")
worksheet = workbook.add_worksheet("Sheet1")
worksheet.write(0, 0, "TITLU")
worksheet.write(0, 1, "CATEGORIE")
worksheet.write(0, 2, "DESCRIERE")
worksheet.write(0, 3, "MONEDA")
worksheet.write(0, 4, "PRET")
worksheet.write(0, 5, "CANTITATE")
worksheet.write(0, 6, "URL_POZA")
row = 1
for car in adauga_excel:
    print(row)
    worksheet.write(row, 0, car[0])
    worksheet.write(row, 1, car[1])
    worksheet.write(row, 2, car[2])
    worksheet.write(row, 3, car[3])
    worksheet.write(row, 4, car[4])
    worksheet.write(row, 5, car[5])
    worksheet.write(row, 6, car[6])
    row += 1
workbook.close()

