from openpyxl import load_workbook
import xlsxwriter

workbook = load_workbook("C:\\Users\\HP\\Desktop\\PieseAuto_ro.xlsx")
worksheet = workbook["Sheet1"]
column_titluri = worksheet["C"]
lista_titluri = [column_titluri[x].value for x in range(len(column_titluri))]

column_preturi = worksheet["E"]
lista_preturi = [column_preturi[x].value for x in range(len(column_preturi))]


workbook_piese = load_workbook("C:\\Users\\HP\\Desktop\\WORK\\RULMENTI\\rulmenti_FAG\\rulmenti_FAG.xlsx")
worksheet_piese = workbook_piese["Sheet1"]
column_piese = worksheet_piese["A"]
piese_mod = [column_piese[x].value for x in range(len(column_piese))]
print(len(piese_mod))
column_piese_set = set(piese_mod)
lista_piese = list(column_piese_set)

lista_cod_pret = []
test = 0
for code in lista_piese:
    code = str(code)
    lista_cod = list(code)
    lista_cod.insert(3, " ")
    lista_cod.insert(8, " ")
    cod_cu_spatiu = "".join(lista_cod)
    for titlu in lista_titluri:
        if str(code) in str(titlu) or str(cod_cu_spatiu) in str(titlu):
            index = lista_titluri.index(titlu)
            pret = lista_preturi[index]
            lista_cod_pret.append([code, pret])
print(lista_cod_pret)


final_workbook = xlsxwriter.Workbook("C:\\Users\\HP\\Desktop\\WORK\\RULMENTI\\rulmenti_FAG\\rulmenti_FAG_pieseauto.xlsx")
final_worksheet = final_workbook.add_worksheet("Sheet1")
row = 0
for code in lista_cod_pret:
    final_worksheet.write(row, 0, code[0])
    final_worksheet.write(row, 1, code[1])
    row += 1

final_workbook.close()